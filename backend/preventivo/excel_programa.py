"""
Generador del PROGRAMA ANUAL DE MANTENIMIENTO PREVENTIVO en Excel.
Replica la hoja "PROGR. MANTTO. EQ." del FORM-DHO-061:
matriz de equipos x meses con doble fila Planificado / Realizado.
"""
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'recursos', 'logo_bateas.jpeg')

THIN = Side(style='thin', color='000000')
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FUENTE = Font(name='Calibri', size=9)
FUENTE_BOLD = Font(name='Calibri', size=9, bold=True)
TITULO = Font(name='Calibri', size=13, bold=True)
GRIS = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
VERDE = PatternFill(start_color='E8F5EE', end_color='E8F5EE', fill_type='solid')
CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
IZQ = Alignment(horizontal='left', vertical='center', wrap_text=True)

MESES = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SET', 'OCT', 'NOV', 'DIC']
CONDICION_SIMBOLO = {'BUENA': '√', 'INTERMEDIA': '~', 'MALA': 'X'}
FRECUENCIA_TEXTO = {'MENSUAL': 'Mensual', 'BIMENSUAL': 'Bimensual', 'TRIMESTRAL': 'Trimestral'}

# Columnas: A=Nº B=Equipo C=Código interno D=Serie E=Marca F=Modelo
#           G=Ubicación H=Condición I=Frecuencia J=(Planif/Realiz) K..V=meses
COL_MES_INICIO = 11  # K


def generar_programa_anual(equipos, anio, subtitulo=''):
    """
    equipos: queryset de Equipo con .programaciones_anio prefetched (del año dado).
    Devuelve BytesIO del xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'PROGR. MANTTO. EQ.'
    construir_hoja_programa(ws, equipos, anio, subtitulo)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def construir_hoja_programa(ws, equipos, anio, subtitulo=''):
    """Llena una hoja de worksheet con la matriz del programa anual."""

    anchos = {'A': 4, 'B': 30, 'C': 14, 'D': 13, 'E': 12, 'F': 12,
              'G': 22, 'H': 10, 'I': 11, 'J': 11}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    for i in range(12):
        ws.column_dimensions[chr(ord('K') + i)].width = 4.5

    # ============ CABECERA ============
    for f in range(1, 5):
        ws.row_dimensions[f].height = 16
    ws.merge_cells('A1:B4')   # logo
    ws.merge_cells('C1:V4')   # título
    c = ws['C1']
    c.value = f'PROGRAMA ANUAL DE MANTENIMIENTO PREVENTIVO DE EQUIPOS - {anio}'
    if subtitulo:
        c.value += f'\n{subtitulo}'
    c.font = TITULO
    c.alignment = CENTRO

    if os.path.exists(LOGO_PATH):
        from PIL import Image as PILImage
        img = PILImage.open(LOGO_PATH)
        img.thumbnail((140, 60))
        buf = io.BytesIO()
        img.convert('RGB').save(buf, format='PNG')
        buf.seek(0)
        logo = XLImage(buf)
        logo.anchor = 'A1'
        ws.add_image(logo)

    # ============ ENCABEZADOS (filas 6-7) ============
    cabeceras = [('A', 'Nº'), ('B', 'Equipo'), ('C', 'Código interno'), ('D', 'Serie'),
                 ('E', 'Marca'), ('F', 'Modelo'), ('G', 'Ubicación'),
                 ('H', 'Condición'), ('I', 'Frecuencia'), ('J', '')]
    for col, texto in cabeceras:
        ws.merge_cells(f'{col}6:{col}7')
        c = ws[f'{col}6']
        c.value = texto
        c.font = FUENTE_BOLD
        c.fill = GRIS
        c.alignment = CENTRO

    ws.merge_cells('K6:V6')
    c = ws['K6']
    c.value = str(anio)
    c.font = FUENTE_BOLD
    c.fill = GRIS
    c.alignment = CENTRO
    for i, mes in enumerate(MESES):
        col = chr(ord('K') + i)
        c = ws[f'{col}7']
        c.value = mes
        c.font = FUENTE_BOLD
        c.fill = GRIS
        c.alignment = CENTRO
    for fila in ws['A6:V7']:
        for celda in fila:
            celda.border = BORDE

    # ============ FILAS DE EQUIPOS (2 filas c/u) ============
    fila = 8
    for idx, eq in enumerate(equipos, start=1):
        seccion = eq.seccion
        ubicacion = f'{seccion.lugar.nombre} - {seccion.nombre}' if seccion else '—'

        # Meses planificados / realizados desde los programas del año
        plan, real = set(), set()
        for p in eq.programaciones_anio:
            plan.add(p.mes_planificado)
            if p.estado == 'EJECUTADO':
                real.add(p.mes_planificado)

        f_plan, f_real = fila, fila + 1

        datos = [('A', idx), ('B', eq.nombre), ('C', eq.codigo_activo),
                 ('D', eq.serie or ''), ('E', eq.marca or ''), ('F', eq.modelo or ''),
                 ('G', ubicacion),
                 ('H', CONDICION_SIMBOLO.get(eq.condicion, '')),
                 ('I', FRECUENCIA_TEXTO.get(eq.frecuencia, eq.frecuencia or ''))]
        for col, valor in datos:
            ws.merge_cells(f'{col}{f_plan}:{col}{f_real}')
            c = ws[f'{col}{f_plan}']
            c.value = valor
            c.font = FUENTE
            c.alignment = IZQ if col in ('B', 'G') else CENTRO

        ws[f'J{f_plan}'] = 'Planificado'
        ws[f'J{f_real}'] = 'Realizado'
        for f in (f_plan, f_real):
            ws[f'J{f}'].font = FUENTE
            ws[f'J{f}'].alignment = IZQ
        ws[f'J{f_real}'].fill = VERDE

        for i in range(12):
            col = chr(ord('K') + i)
            mes = i + 1
            cp = ws[f'{col}{f_plan}']
            cr = ws[f'{col}{f_real}']
            cp.value = 'X' if mes in plan else None
            cr.value = 'X' if mes in real else None
            cp.font = FUENTE_BOLD
            cr.font = FUENTE_BOLD
            cp.alignment = CENTRO
            cr.alignment = CENTRO
            cr.fill = VERDE

        for fx in ws[f'A{f_plan}:V{f_real}']:
            for celda in fx:
                celda.border = BORDE

        fila += 2
