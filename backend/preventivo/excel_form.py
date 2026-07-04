"""
Generador del FORM-DHO-061 en Excel.
Replica los cuadros, códigos y formalidades del formato oficial
"INFORME DE TRABAJOS DE MANTENIMIENTO DE EQUIPOS Y CAMPAMENTO"
y lo llena con los datos reales del informe.
"""
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'recursos', 'logo_bateas.jpeg')

# ---- Estilos base ----
THIN = Side(style='thin', color='000000')
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FUENTE = Font(name='Calibri', size=10)
FUENTE_BOLD = Font(name='Calibri', size=10, bold=True)
TITULO = Font(name='Calibri', size=12, bold=True)
GRIS = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
IZQ = Alignment(horizontal='left', vertical='center', wrap_text=True)
IZQ_ARRIBA = Alignment(horizontal='left', vertical='top', wrap_text=True)

SCORE_TEXTO = {1: 'Excelente', 2: 'Bueno', 3: 'Regular', 4: 'Malo', 5: 'Falla'}


def _bordear(ws, rango):
    for fila in ws[rango]:
        for celda in fila:
            celda.border = BORDE


def _titulo_seccion(ws, fila, texto):
    ws.merge_cells(f'B{fila}:H{fila}')
    c = ws[f'B{fila}']
    c.value = texto
    c.font = FUENTE_BOLD
    c.fill = GRIS
    c.alignment = CENTRO
    _bordear(ws, f'B{fila}:H{fila}')


def _campo(ws, celda_label, celda_valor_ini, celda_valor_fin, label, valor):
    ws[celda_label] = label
    ws[celda_label].font = FUENTE_BOLD
    ws[celda_label].alignment = IZQ
    if celda_valor_ini != celda_valor_fin:
        ws.merge_cells(f'{celda_valor_ini}:{celda_valor_fin}')
    c = ws[celda_valor_ini]
    c.value = valor or '—'
    c.font = FUENTE
    c.alignment = IZQ


def _imagen_ajustada(archivo_o_buffer, ancho_px, alto_px):
    """Redimensiona con PIL respetando proporción y devuelve XLImage."""
    from PIL import Image as PILImage
    img = PILImage.open(archivo_o_buffer)
    img.thumbnail((ancho_px, alto_px))
    buf = io.BytesIO()
    img.convert('RGB').save(buf, format='PNG')
    buf.seek(0)
    xl = XLImage(buf)
    return xl


def _firma_imagen(usuario, ancho=170, alto=70):
    """Firma desde trazos normalizados -> XLImage (o None)."""
    if not usuario or not getattr(usuario, 'firma_trazos', None):
        return None
    try:
        from PIL import Image as PILImage, ImageDraw
        w, h = 340, 140
        img = PILImage.new('RGB', (w, h), (255, 255, 255))
        draw = ImageDraw.Draw(img)
        for trazo in usuario.firma_trazos:
            puntos = [(round(x * w), round(y * h)) for x, y in trazo]
            if len(puntos) >= 2:
                draw.line(puntos, fill=(26, 29, 35), width=3, joint='curve')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        xl = XLImage(buf)
        xl.width, xl.height = ancho, alto
        return xl
    except Exception:
        return None


def generar_form_dho_061(informe):
    """Construye el workbook FORM-DHO-061 y devuelve un BytesIO listo para descargar."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'FORM-DHO-061'
    construir_hoja_informe(ws, informe)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def construir_hoja_informe(ws, informe):
    """Llena una hoja de worksheet con el FORM-DHO-061 del informe dado."""
    equipo = informe.programa.equipo if informe.programa else None
    seccion = equipo.seccion if equipo else None
    detalles = list(informe.detalles_score.select_related('componente').all())

    # Columnas B..H (A queda como margen, igual que el original)
    anchos = {'A': 2, 'B': 24, 'C': 13, 'D': 13, 'E': 13, 'F': 17, 'G': 12, 'H': 15}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    # ================= CABECERA (filas 2-5) =================
    for f in range(2, 6):
        ws.row_dimensions[f].height = 18

    ws.merge_cells('B2:B5')       # logo
    ws.merge_cells('C2:F5')       # título
    c = ws['C2']
    c.value = 'INFORME DE TRABAJOS DE MANTENIMIENTO DE EQUIPOS Y CAMPAMENTO'
    c.font = TITULO
    c.alignment = CENTRO

    codigos = [('G2', 'Código:', 'H2', 'FORM-DHO-061'),
               ('G3', 'Versión:', 'H3', '1'),
               ('G4', 'Fecha:', 'H4', informe.fecha.strftime('%d.%m.%Y')),
               ('G5', 'Pagina:', 'H5', '1 de 1')]
    for cl, lab, cv, val in codigos:
        ws[cl] = lab
        ws[cl].font = FUENTE_BOLD
        ws[cl].alignment = IZQ
        ws[cv] = val
        ws[cv].font = FUENTE
        ws[cv].alignment = IZQ
    _bordear(ws, 'B2:H5')

    if os.path.exists(LOGO_PATH):
        logo = _imagen_ajustada(LOGO_PATH, 150, 68)
        logo.anchor = 'B2'
        ws.add_image(logo)

    # ================= DATOS GENERALES (filas 7-11) =================
    zona = seccion.lugar.zona.nombre if seccion else None
    lugar_ubic = f'{seccion.lugar.nombre} - {seccion.nombre}' if seccion else None
    marca_serie = ' / '.join(x for x in [equipo.marca if equipo else None,
                                         equipo.serie if equipo else None] if x) or None
    responsables = informe.tecnico.get_full_name() or informe.tecnico.username

    _campo(ws, 'B7', 'C7', 'D7', 'Area:', zona)
    _campo(ws, 'F7', 'G7', 'H7', 'Fecha:', informe.fecha.strftime('%d.%m.%Y'))
    _campo(ws, 'B8', 'C8', 'D8', 'Lugar / Ubicación:', lugar_ubic)
    _campo(ws, 'F8', 'G8', 'H8', 'Hora de Inicio:',
           informe.hora_inicio.strftime('%H:%M') if informe.hora_inicio else None)
    _campo(ws, 'B9', 'C9', 'D9', 'Equipo / Otros:', equipo.nombre if equipo else None)
    _campo(ws, 'F9', 'G9', 'H9', 'Hora de Termino:',
           informe.hora_fin.strftime('%H:%M') if informe.hora_fin else None)
    _campo(ws, 'B10', 'C10', 'D10', 'Tipo de Mantenimiento:', 'Preventivo')
    _campo(ws, 'F10', 'G10', 'H10', 'Marca / Serie:', marca_serie)
    _campo(ws, 'B11', 'C11', 'H11', 'Responsables del Trabajo:', responsables)
    _bordear(ws, 'B7:H11')

    fila = 13

    # ================= HALLAZGO / REPORTE =================
    _titulo_seccion(ws, fila, 'HALLAZGO / REPORTE / DESVIACIÓN / OBSERVACIÓN')
    fila += 1
    fin_hallazgo = fila + 3
    ws.merge_cells(f'B{fila}:H{fin_hallazgo}')
    c = ws[f'B{fila}']
    c.value = informe.hallazgos_generales or ''
    c.font = FUENTE
    c.alignment = IZQ_ARRIBA
    _bordear(ws, f'B{fila}:H{fin_hallazgo}')
    fila = fin_hallazgo + 2

    # ================= ESTADO ACTUAL DEL EQUIPO (evaluación) =================
    _titulo_seccion(ws, fila, 'ESTADO ACTUAL DEL EQUIPO')
    fila += 1

    headers = [('B', 'Componente'), ('C', 'Encontrado'), ('D', 'Intervención'),
               ('E', 'Final'), ('F', 'Detalle / Observación')]
    ws.merge_cells(f'F{fila}:H{fila}')
    for col, texto in headers:
        c = ws[f'{col}{fila}']
        c.value = texto
        c.font = FUENTE_BOLD
        c.fill = GRIS
        c.alignment = CENTRO
    _bordear(ws, f'B{fila}:H{fila}')
    fila += 1

    if detalles:
        for d in detalles:
            ini = d.score_inicial or d.score_valor
            ws[f'B{fila}'] = d.componente.nombre_componente
            ws[f'C{fila}'] = f'{ini} - {SCORE_TEXTO.get(ini, "")}'
            ws[f'D{fila}'] = 'Sí' if d.intervencion else 'No'
            ws[f'E{fila}'] = f'{d.score_valor} - {SCORE_TEXTO.get(d.score_valor, "")}'
            detalle = ' | '.join(x for x in [d.detalle_intervencion, d.observacion_tecnica] if x)
            if d.derivado:
                detalle = (detalle + ' | ' if detalle else '') + 'DERIVADO A CORRECTIVO' + (' (TERCERO)' if d.requiere_tercero else '')
            ws.merge_cells(f'F{fila}:H{fila}')
            ws[f'F{fila}'] = detalle
            for col in 'BCDEF':
                ws[f'{col}{fila}'].font = FUENTE
                ws[f'{col}{fila}'].alignment = IZQ if col in ('B', 'F') else CENTRO
            _bordear(ws, f'B{fila}:H{fila}')
            fila += 1
    else:
        ws.merge_cells(f'B{fila}:H{fila}')
        ws[f'B{fila}'] = 'Sin componentes evaluados.'
        ws[f'B{fila}'].font = FUENTE
        _bordear(ws, f'B{fila}:H{fila}')
        fila += 1

    fila += 1

    # ================= EVIDENCIAS =================
    _titulo_seccion(ws, fila, 'EVIDENCIAS')
    fila += 1

    ws.merge_cells(f'B{fila}:C{fila}')
    ws.merge_cells(f'D{fila}:E{fila}')
    ws.merge_cells(f'F{fila}:H{fila}')
    for col, texto in [('B', 'ANTES'), ('D', 'DESPUES'), ('F', 'DESCRIPCIÓN')]:
        c = ws[f'{col}{fila}']
        c.value = texto
        c.font = FUENTE_BOLD
        c.fill = GRIS
        c.alignment = CENTRO
    _bordear(ws, f'B{fila}:H{fila}')
    fila += 1

    # Bloque de fotos (alto fijo, como el cuadro original)
    ini_fotos = fila
    fin_fotos = fila + 11
    for f in range(ini_fotos, fin_fotos + 1):
        ws.row_dimensions[f].height = 16
    ws.merge_cells(f'B{ini_fotos}:C{fin_fotos}')
    ws.merge_cells(f'D{ini_fotos}:E{fin_fotos}')
    ws.merge_cells(f'F{ini_fotos}:H{fin_fotos}')

    # Descripción del cuadro: intervenciones + pendientes derivados
    intervenciones = [f'{d.componente.nombre_componente}: {d.detalle_intervencion}'
                      for d in detalles if d.intervencion and d.detalle_intervencion]
    pendientes = [d.componente.nombre_componente for d in detalles if d.derivado]
    texto_desc = ''
    if intervenciones:
        texto_desc += 'Trabajos realizados:\n' + '\n'.join(f'- {t}' for t in intervenciones)
    if pendientes:
        texto_desc += ('\n\n' if texto_desc else '') + 'Pendientes (derivados a correctivo):\n' + '\n'.join(f'- {p}' for p in pendientes)
    c = ws[f'F{ini_fotos}']
    c.value = texto_desc
    c.font = FUENTE
    c.alignment = IZQ_ARRIBA
    _bordear(ws, f'B{ini_fotos}:H{fin_fotos}')

    # Insertar la primera foto de cada tipo dentro de su cuadro
    for tipo, col in [('ANTES', 'B'), ('DESPUES', 'D')]:
        ev = informe.evidencias.filter(tipo=tipo).first()
        if ev:
            try:
                img = _imagen_ajustada(ev.foto.path, 175, 165)
                img.anchor = f'{col}{ini_fotos}'
                ws.add_image(img)
            except Exception:
                pass

    fila = fin_fotos + 3

    # ================= FIRMAS =================
    ini_firma = fila
    for f in range(ini_firma, ini_firma + 4):
        ws.row_dimensions[f].height = 16

    ws.merge_cells(f'B{ini_firma}:C{ini_firma + 3}')
    ws.merge_cells(f'F{ini_firma}:H{ini_firma + 3}')
    _bordear(ws, f'B{ini_firma}:C{ini_firma + 3}')
    _bordear(ws, f'F{ini_firma}:H{ini_firma + 3}')

    firma_tec = _firma_imagen(informe.tecnico)
    if firma_tec:
        firma_tec.anchor = f'B{ini_firma}'
        ws.add_image(firma_tec)
    firma_sup = _firma_imagen(informe.supervisor)
    if firma_sup:
        firma_sup.anchor = f'F{ini_firma}'
        ws.add_image(firma_sup)

    fila_label = ini_firma + 4
    ws.merge_cells(f'B{fila_label}:C{fila_label}')
    ws.merge_cells(f'F{fila_label}:H{fila_label}')
    nombre_sup = ''
    if informe.supervisor:
        nombre_sup = informe.supervisor.get_full_name() or informe.supervisor.username
    ws[f'B{fila_label}'] = f'Firma del Téc. Mantenimiento\n{responsables}'
    ws[f'F{fila_label}'] = f'Firma del Supervisor\n{nombre_sup}'
    for col in ('B', 'F'):
        c = ws[f'{col}{fila_label}']
        c.font = FUENTE_BOLD
        c.alignment = CENTRO
    ws.row_dimensions[fila_label].height = 30
