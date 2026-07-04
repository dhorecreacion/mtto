# ============================================================
# Carga masiva de equipos vía Excel
#   - generar_plantilla_equipos(): plantilla con desplegables y valores válidos
#   - importar_equipos(): lee el archivo llenado y crea los equipos
# ============================================================
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Equipo, Seccion, MantenimientoComponente

# Columnas de la plantilla: (encabezado, ancho)
COLUMNAS = [
    ('CODIGO ACTIVO *',            18),
    ('NOMBRE *',                   30),
    ('SERIE',                      16),
    ('MARCA',                      16),
    ('MODELO',                     16),
    ('SECCION (Lugar - Sección)',  34),
    ('CONDICION *',                14),
    ('CRITICIDAD',                 13),
    ('FRECUENCIA',                 14),
    ('CATEGORIA',                  24),
    ('ESTADO OPERATIVO',           18),
    ('COMPONENTES (separados por ;)', 40),
    ('OBSERVACIONES',              40),
]
FILAS_PLANTILLA = 300  # filas con desplegables listas para llenar

FUENTE_BOLD = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
AZUL = PatternFill('solid', fgColor='036494')
GRIS = PatternFill('solid', fgColor='F2F2F2')
CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
BORDE = Border(*[Side(style='thin')] * 4)


def _mapa_choices(choices):
    """{'BUENA': 'BUENA', 'buena': 'BUENA', ...}: acepta el valor o la etiqueta, sin distinguir mayúsculas."""
    mapa = {}
    for valor, etiqueta in choices:
        mapa[valor.lower()] = valor
        mapa[etiqueta.lower()] = valor
    return mapa


def _secciones_activas():
    return Seccion.objects.filter(activo=True).select_related('lugar').order_by('lugar__nombre', 'nombre')


def _etiqueta_seccion(s):
    return f'{s.lugar.nombre} - {s.nombre}'


def generar_plantilla_equipos():
    wb = Workbook()

    # ---- Hoja VALORES (catálogos que alimentan los desplegables) ----
    ws_val = wb.create_sheet('VALORES')
    secciones = list(_secciones_activas())
    catalogos = [
        ('SECCION',    [_etiqueta_seccion(s) for s in secciones]),
        ('CONDICION',  [et for _, et in Equipo.Condicion.choices]),
        ('CRITICIDAD', [et for _, et in Equipo.Criticidad.choices]),
        ('FRECUENCIA', [et for _, et in Equipo.Frecuencia.choices]),
        ('CATEGORIA',  [et for _, et in Equipo.Categoria.choices]),
        ('ESTADO',     [et for _, et in Equipo.EstadoOperativo.choices]),
    ]
    rangos = {}
    for col, (titulo, valores) in enumerate(catalogos, start=1):
        letra = get_column_letter(col)
        c = ws_val.cell(row=1, column=col, value=titulo)
        c.font = Font(bold=True)
        for fila, v in enumerate(valores, start=2):
            ws_val.cell(row=fila, column=col, value=v)
        ws_val.column_dimensions[letra].width = 34 if titulo == 'SECCION' else 22
        if valores:
            rangos[titulo] = f'VALORES!${letra}$2:${letra}${len(valores) + 1}'

    # ---- Hoja EQUIPOS (la que se llena) ----
    ws = wb.active
    ws.title = 'EQUIPOS'
    ws.freeze_panes = 'A2'
    for col, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = FUENTE_BOLD
        c.fill = AZUL
        c.alignment = CENTRO
        c.border = BORDE
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[1].height = 30

    # Desplegables: columna de la plantilla -> catálogo
    desplegables = {6: 'SECCION', 7: 'CONDICION', 8: 'CRITICIDAD',
                    9: 'FRECUENCIA', 10: 'CATEGORIA', 11: 'ESTADO'}
    for col, titulo in desplegables.items():
        if titulo not in rangos:
            continue
        dv = DataValidation(type='list', formula1=rangos[titulo], allow_blank=True,
                            errorTitle='Valor inválido',
                            error='Elige un valor de la lista (hoja VALORES).')
        ws.add_data_validation(dv)
        letra = get_column_letter(col)
        dv.add(f'{letra}2:{letra}{FILAS_PLANTILLA + 1}')

    # ---- Hoja de instrucciones ----
    ws_ins = wb.create_sheet('INSTRUCCIONES')
    ws_ins.column_dimensions['A'].width = 100
    instrucciones = [
        'PLANTILLA DE CARGA MASIVA DE EQUIPOS',
        '',
        '1. Llena la hoja EQUIPOS, una fila por equipo. Las columnas con * son obligatorias.',
        '2. Usa los desplegables (o copia los valores exactos de la hoja VALORES).',
        '3. SECCION: elige del desplegable con formato "Lugar - Sección". Déjala vacía si el equipo está en almacén.',
        '4. FRECUENCIA: solo aplica a equipos de categoría Industrial (para su programa preventivo).',
        '5. COMPONENTES: nombres separados por punto y coma. Ej: Motor; Quemador; Panel de control',
        '6. Si no llenas CRITICIDAD, CATEGORIA o ESTADO OPERATIVO, se asume: Media, Industrial, En Uso.',
        '7. Los códigos de activo que ya existan en el sistema se omiten (no se duplican ni se modifican).',
        '8. Guarda el archivo y súbelo en Gestión → Equipos → Importar Excel.',
    ]
    for fila, texto in enumerate(instrucciones, start=1):
        c = ws_ins.cell(row=fila, column=1, value=texto)
        if fila == 1:
            c.font = Font(bold=True, size=12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf


def importar_equipos(archivo, usuario):
    """
    Lee la hoja EQUIPOS de la plantilla y crea los equipos (con sus componentes).
    Devuelve: {'creados': int, 'omitidos': [códigos ya existentes], 'errores': [{'fila', 'error'}]}
    Las filas con error se saltan; el resto sí se carga.
    """
    wb = load_workbook(archivo, data_only=True)
    ws = wb['EQUIPOS'] if 'EQUIPOS' in wb.sheetnames else wb.active

    map_condicion  = _mapa_choices(Equipo.Condicion.choices)
    map_criticidad = _mapa_choices(Equipo.Criticidad.choices)
    map_frecuencia = _mapa_choices(Equipo.Frecuencia.choices)
    map_categoria  = _mapa_choices(Equipo.Categoria.choices)
    map_estado     = _mapa_choices(Equipo.EstadoOperativo.choices)

    # Secciones por etiqueta "Lugar - Sección" y por nombre a secas (si es único)
    secciones = {}
    por_nombre = {}
    for s in _secciones_activas():
        secciones[_etiqueta_seccion(s).lower()] = s
        por_nombre.setdefault(s.nombre.lower(), []).append(s)
    for nombre, lista in por_nombre.items():
        if len(lista) == 1 and nombre not in secciones:
            secciones[nombre] = lista[0]

    codigos_existentes = set(
        Equipo.objects.values_list('codigo_activo', flat=True)
    )

    creados, omitidos, errores = 0, [], []
    vistos_en_archivo = set()

    for idx, fila in enumerate(ws.iter_rows(min_row=2, max_col=len(COLUMNAS), values_only=True), start=2):
        celdas = [str(v).strip() if v is not None else '' for v in fila]
        if not any(celdas):
            continue  # fila vacía
        (codigo, nombre, serie, marca, modelo, seccion_txt, condicion_txt,
         criticidad_txt, frecuencia_txt, categoria_txt, estado_txt,
         componentes_txt, observaciones) = celdas

        problemas = []
        if not codigo:
            problemas.append('falta el código de activo')
        if not nombre:
            problemas.append('falta el nombre')

        if codigo and codigo in codigos_existentes:
            omitidos.append(codigo)
            continue
        if codigo and codigo.lower() in vistos_en_archivo:
            problemas.append(f'código "{codigo}" repetido en el archivo')

        seccion = None
        if seccion_txt:
            seccion = secciones.get(seccion_txt.lower())
            if seccion is None:
                problemas.append(f'sección "{seccion_txt}" no existe (usa el formato "Lugar - Sección")')

        def resolver(texto, mapa, campo, obligatorio=False, defecto=None):
            if not texto:
                if obligatorio:
                    problemas.append(f'falta {campo}')
                return defecto
            valor = mapa.get(texto.lower())
            if valor is None:
                problemas.append(f'{campo} "{texto}" inválido')
            return valor

        condicion  = resolver(condicion_txt,  map_condicion,  'la condición', obligatorio=True)
        criticidad = resolver(criticidad_txt, map_criticidad, 'la criticidad', defecto=Equipo.Criticidad.MEDIA)
        frecuencia = resolver(frecuencia_txt, map_frecuencia, 'la frecuencia', defecto=None)
        categoria  = resolver(categoria_txt,  map_categoria,  'la categoría', defecto=Equipo.Categoria.INDUSTRIAL)
        estado     = resolver(estado_txt,     map_estado,     'el estado operativo', defecto=Equipo.EstadoOperativo.EN_USO)

        if problemas:
            errores.append({'fila': idx, 'error': '; '.join(problemas)})
            continue

        equipo = Equipo.objects.create(
            codigo_activo=codigo, nombre=nombre,
            serie=serie or None, marca=marca or None, modelo=modelo or None,
            seccion=seccion, condicion=condicion, criticidad=criticidad,
            frecuencia=frecuencia, categoria_mantenimiento=categoria,
            estado_operativo=estado, observaciones=observaciones or None,
            creado_por=usuario,
        )
        for nombre_comp in filter(None, (c.strip() for c in componentes_txt.split(';'))):
            MantenimientoComponente.objects.create(
                equipo=equipo, nombre_componente=nombre_comp[:100], creado_por=usuario,
            )
        vistos_en_archivo.add(codigo.lower())
        creados += 1

    return {'creados': creados, 'omitidos': omitidos, 'errores': errores}
